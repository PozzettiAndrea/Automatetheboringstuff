import os
import openpyxl

workbook = openpyxl.load_workbook("example.xlsx")
print("Sheet names:","\n")

print(workbook.get_sheet_names(),"\n")

print("Selecting Sheet1 by name, printing cell A1 and cellA1.value:","\n")

sheet = workbook.get_sheet_by_name("Sheet1")
print(sheet["A1"],sheet["A1"].value,"\n")

print("B1 and C1 values:","\n")

print(sheet["B1"].value, sheet["C1"].value,"\n")
print("With sheet.cell method you can select by row and column: example B1 row 1 column 2","\n", "print(sheet.cell(row=1, column = 2)).value","\n")
print(sheet.cell(row=1, column = 2).value)

os.system("pause")